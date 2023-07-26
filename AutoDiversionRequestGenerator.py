from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import os

#============================= START USER INPUT ================================

DIVERSION_REQUEST_STARTING_NUMBER = 42
CONTRACT_CODE = "S-48012"
DATE_SUBMITTED = "6-7-23" # USE MM-DD-YY FORMAT
STARTING_1A = 1
STARTING_3A = 1
STARTING_6A = 1
STARTING_7A = 1
STARTING_5A = 1
STARTING_1C = 1
STARTING_3C = 1

#============================== END USER INPUT =================================
#open excel file
workbook = load_workbook(filename="C:\\Users\\mdas\\Documents\\PythonProject\\AutoDiversionRequrestGenerator\\Planned_Work.xlsx",  data_only=True)
#select sheet name
sheet = workbook["2024 Weeknight GO"]

#global vars
alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

stops_g_line = ['Church', 'Beverly Road', 'Cortelyou Road', 'Newkirk Plaza', '7th', '15th Street-Prospect Park', 
                   'Fort Hamilton Parkway', '4th-9th Street', 'Smith-9th Street', '7th', '15th Street-Prospect Park', 
                   'Fort Hamilton Parkway', 'Church', '15th Street-Prospect Park', '7th', '4th-9th Street', '9th Street', 
                   'Smith-9th Street', 'Carroll Street', 'Bergen Street', 'Hoyt-Schermerhorn Street', 'Lafayette', 
                   'Fulton Street', 'Clinton-Washington', 'Classon', 'Bedford-Nostrand', 'Metropolitan', 'Broadway', 'Flushing', 
                   'Myrtle-Willoughby', 'Bedford-Nostrand', 'Lorimer Street', 'Metropolitan', 'Nassau', 'Greenpoint', '21st Street', 
                   'Court Square']

directory_path = r'C:\\Users\\mdas\\Documents\\PythonProject\\AutoDiversionRequrestGenerator\\Requests'
os.makedirs(directory_path, exist_ok=True)
os.chmod(directory_path, 0o755)  # Adjust the permission value as needed


#store potential column headers, eventually will be stored in JSON file as list expands
week = ["Week","Week #"]
divNum = ["Diversion","Diversion #"]
sect = ["Section","Label ACC"]
start = ["Start Date","Label ADD"]
end = ["End Date","Label AEE"]
divLim = ["Diversion Limits","Label AHH"]
allLabels = week + divNum + sect + start + end + divLim

weekCol = -1
divNumCol = -1
sectCol = -1
startCol = -1
endCol = -1
divLimCol = -1

#search for title column
def search_excel_diagonal(search_value, min_layer, max_layer):
  rows = sheet.max_row
  cols = sheet.max_column

  for layer in range(min_layer, max_layer + 1):
      for row in range(1, rows + 1):
          for col in range(1, cols + 1):
              if row + col == layer:
                  cell_value = sheet.cell(row=row, column=col).value
                  if cell_value in search_value:
                      return col,row
                      #return sheet.cell(row=row, column=col).coordinate
  
  # If the search value is not found within the maximum layer, return -1
  return -1

#stores the value of the top-left corner of the excel table

#print(tableHeaderOrigin)
def assignColCoors():
  global weekCol, divNumCol, sectCol, startCol, endCol, divLimCol
  
  row = tableHeaderOrigin[1]
  col = tableHeaderOrigin[0]
  for i in range(col,100):
    if weekCol != -1 and divNumCol != -1 and sectCol != -1 and startCol != -1 and endCol != -1 and divLimCol != -1:
      break
    else:
      if str(sheet.cell(row=row, column=i).value).strip() in week:
        weekCol = i
      elif str(sheet.cell(row=row, column=i).value).strip() in divNum:
        divNumCol = i
      elif str(sheet.cell(row=row, column=i).value).strip() in sect:
        sectCol = i
      elif str(sheet.cell(row=row, column=i).value).strip() in start:
        startCol = i
      elif str(sheet.cell(row=row, column=i).value).strip() in end:
        endCol = i
      elif str(sheet.cell(row=row, column=i).value).strip() in divLim:
        divLimCol = i

def IterTable(ReqNum):
  currentRow = tableHeaderOrigin[1] + 1
  count = 0
  addition1A = 0
  addition3A = 0
  addition6A = 0
  addition7A = 0
  addition5A = 0
  while(str(sheet.cell(row=currentRow,column=weekCol).value) != "None"):
    CurrWeek = str(sheet.cell(row=currentRow,column=weekCol).value)
    CurrDivNum = str(sheet.cell(row=currentRow,column=divNumCol).value)
    CurrSect = str(sheet.cell(row=currentRow,column=sectCol).value)
    CurrStart = str(sheet.cell(row=currentRow,column=startCol).value)
    CurrEnd = str(sheet.cell(row=currentRow,column=endCol).value)
    CurrDivLim = str(sheet.cell(row=currentRow,column=divLimCol).value)
    if CurrDivNum == "1A":
      listNumDiv = STARTING_1A + addition1A
      addition1A += 4
    elif CurrDivNum == "3A":
      listNumDiv = STARTING_3A + addition3A
      addition3A += 4
    elif CurrDivNum == "6A":
      listNumDiv = STARTING_6A + addition6A
      addition6A += 4
    elif CurrDivNum == "7A":
      listNumDiv = STARTING_7A + addition7A
      addition7A += 4
    elif CurrDivNum == "5A":
      listNumDiv = STARTING_5A + addition5A
      addition5A += 4
    else:
      listNumDiv = "ERR"
    WriteOutDivReq(ReqNum + count,CurrWeek,CurrDivNum,CurrSect,CurrStart,CurrEnd,CurrDivLim,listNumDiv)
    currentRow += 1
    count += 1

def LimToTrack(limits):
  ans = []
  for i in range(len(limits)):
    limits[i] = limits[i].replace("/","-").split("-")
    for j in range(len(limits[i])):
      if not(limits[i][j].isdigit()) and limits[i][j] != "":
        ans.append(limits[i][j])
  ans = list(set(ans))    
  return str(ans).replace("[","").replace("]","").replace("'","")

def LimToStations(limits):
    mentioned_stations = []
    indices = []

    for station in stops_g_line:
        index = limits.find(station)
        if index != -1:  # If station found in limits
            mentioned_stations.append(station)
            indices.append(index)

    # Sort mentioned_stations based on the indices to preserve order of appearance
    mentioned_stations = [station for _, station in sorted(zip(indices, mentioned_stations))]

    return list(set(mentioned_stations))
#    
def WriteOutDivReq(num,w,dn,se,st,e,dl,lnd):
  fileName = generateFileName(num,st,e,dn,lnd)
  if fileName != None:
    limits = findDivLimits(dl)
    
    print(fileName)
    fileName = "C:\\Users\\mdas\\Documents\\PythonProject\\AutoDiversionRequrestGenerator\\Requests\\" + fileName + ".xlsx"
    workbook2 = load_workbook(filename="C:\\Users\\mdas\\Documents\\PythonProject\\AutoDiversionRequrestGenerator\\DivRequest.xlsx")

    source_sheet = workbook2.active
    new_sheet = workbook2.copy_worksheet(source_sheet)
    new_sheet.title = 'Sheet1'
    workbook2.save(fileName)
    #at this point, workbook is loaded to template
    workbook2 = load_workbook(filename=fileName)
    source_sheet = workbook2.active
    source_sheet.cell(row=8,column=45).value = "0" + str(num)
    source_sheet.cell(row=4,column=44).value = str(w)

    source_sheet.cell(row=13,column=17).value = limits[0] #start limits
    source_sheet.cell(row=13,column=32).value = limits[1] #end limits

    source_sheet.cell(row=11,column=55).value = str(st.replace("00:00:00","").strip().split("-")[1]) # month
    source_sheet.cell(row=11,column=58).value = str(st.replace("00:00:00","").strip().split("-")[2]) # day
    source_sheet.cell(row=11,column=61).value = str(st.replace("00:00:00","").strip().split("-")[0]) # year
    source_sheet.cell(row=11,column=66).value = str(e.replace("00:00:00","").strip().split("-")[1]) # month
    source_sheet.cell(row=11,column=69).value = str(e.replace("00:00:00","").strip().split("-")[2]) # day
    source_sheet.cell(row=11,column=72).value = str(e.replace("00:00:00","").strip().split("-")[0]) # year

    source_sheet.cell(row=10,column=25).value = "TRACK(s) O/S: " + LimToTrack(limits) #Tracks
    stations = LimToStations(dl)
    source_sheet.cell(row=20,column=17).value = LimToStations(dl)[0]

    ans = ""
    for i in range(lnd,lnd + 3):
      ans += str(i) + ","
    ans = "Div. " + dn + "-" + ans + str(lnd + 3)

    source_sheet.cell(row=32,column=1).value = ans

    if len(stations) > 1:
      source_sheet.cell(row=20,column=38).value = LimToStations(dl)[1]
    workbook2.save(fileName)

#returns a list where [0] is the start and [1] is the end
def findDivLimits(comment):
  array = comment.split()
  ans = []
  for i in range(len(array)):
    if array[i][0] in alphabet and array[i][1].isdigit() and array[i][2] == "-":
      ans.append(array[i])
  return ans

def generateFileName(divNum,st,e,designation,designationNumber):
  if designationNumber == "ERR":
    return None
  if divNum < 100:
    stringNum = "0" + str(divNum)
  else:
    stringNum = str(divNum)

  date = st.replace("00:00:00","").strip()
  dateEnd = e.replace("00:00:00","").strip()
  dateL = date.split("-")
  dateEndL = dateEnd.split("-")
  date = str(int(dateL[1])) + "." + str(int(dateL[2])) + "." + dateL[0][-2:]
  dateEnd = str(int(dateEndL[1])) + "." + str(int(dateEndL[2])) + "." + dateEndL[0][-2:]
  ans = ""
  for i in range(designationNumber,designationNumber + 3):
    ans += str(i) + ","
  ans = "(" + designation + "-" + ans + str(designationNumber + 3) + ")"
  
  name = str(CONTRACT_CODE) + " - " + "DR-" + str(stringNum) + " - " + str(date) + "-" + str(dateEnd) + " " + str(ans)
  
  return name

#======================= CODE EXECUTION START =================================

tableHeaderOrigin = search_excel_diagonal(allLabels,0,15) #search table origin

assignColCoors() # assign weekCol - divLimCol (line 20)

IterTable(DIVERSION_REQUEST_STARTING_NUMBER)